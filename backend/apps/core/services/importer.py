"""
Unified Bulk Data Import & Migration Engine for ShoeDZ.
Supports Products & Variants (with opening stock movements), Suppliers, and Clients.
Provides multi-lingual auto-mapping (FR, EN, AR), dry-run validation, duplicate handling,
and dynamic localized template generation (.xlsx and .csv).
"""
import csv
import io
import re
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.db import transaction
from django.utils.translation import gettext_lazy as _

from apps.core.models import Branch, Tenant, WILAYA_CHOICES
from apps.inventory.models import (
    CategoryChoices,
    GenderChoices,
    MovementReasonChoices,
    Product,
    SeasonChoices,
    StockMovement,
    Variant,
)
from apps.suppliers.models import Supplier
from apps.clients.models import Client, ClientTypeChoices


# ─────────────────────────────────────────────────────────────────────────────
# Multi-Lingual Column Synonym Dictionaries (FR, EN, AR)
# ─────────────────────────────────────────────────────────────────────────────

PRODUCT_SYNONYMS: Dict[str, List[str]] = {
    "name": [
        "name", "nom", "designation", "désignation", "libelle", "libellé",
        "article", "produit", "product", "item", "title",
        "الاسم", "اسم المنتج", "التعيين", "المادة", "اسم الموديل", "اسم الحذاء", "السلعة"
    ],
    "brand": [
        "brand", "marque", "fabricant", "label", "fournisseur_marque",
        "الماركة", "العلامة", "العلامة التجارية", "الشركة"
    ],
    "reference": [
        "reference", "référence", "ref", "sku", "code_article", "code",
        "model", "modèle", "code_modele", "ref_fournisseur",
        "المرجع", "رمز الموديل", "رمز المادة", "كود المنتج", "الرمز"
    ],
    "category": [
        "category", "catégorie", "type", "famille", "rayon", "genre_produit",
        "التصنيف", "القسم", "الفئة", "النوع", "الصنف"
    ],
    "gender": [
        "gender", "genre", "sexe", "public", "cible",
        "الجنس", "الفئة المستهدفة", "الفئة"
    ],
    "season": [
        "season", "saison", "collection",
        "الموسم", "الفصل"
    ],
    "purchase_price": [
        "purchase_price", "prix_achat", "prix_d_achat", "prix achat", "cout", "coût",
        "cost", "pamp", "buying_price", "prix d'achat",
        "سعر الشراء", "سعر التكلفة", "تكلفة الشراء", "ثمن الشراء", "سعر الشراء دج"
    ],
    "sale_price": [
        "sale_price", "selling_price", "prix_vente", "prix_de_vente", "prix vente",
        "prix", "price", "pv", "prix de vente", "prix_detail", "prix détail",
        "سعر البيع", "سعر البيع بالتجزئة", "الثمن", "السعر", "سعر التجزئة", "سعر البيع دج"
    ],
    "wholesale_price": [
        "wholesale_price", "prix_gros", "prix_de_gros", "prix gros", "pvg",
        "سعر الجملة", "سعر البيع بالجملة", "ثمن الجملة", "سعر الجملة دج"
    ],
    "pairs_per_carton": [
        "pairs_per_carton", "carton", "colisage", "paquet", "paires_par_carton",
        "pcs_per_box", "pairs_box", "colis",
        "عدد الأزواج في الكرتون", "كرتون", "حزمة", "الأزواج في الكرتون", "العبوة"
    ],
    "alert_threshold": [
        "alert_threshold", "seuil_alerte", "min_stock", "stock_min", "alerte",
        "seuil", "min_alert", "alerte_stock",
        "حد التنبيه", "الحد الأدنى للكمية", "تنبيه المخزون", "الحد الأدنى"
    ],
    "size_eu": [
        "size", "size_eu", "pointure", "taille", "shoe_size", "eu_size", "pointures",
        "المقاس", "القياس", "الحجم", "النمرة", "مقاس الحذاء"
    ],
    "colour": [
        "colour", "color", "couleur", "coloris", "teinte",
        "اللون", "لون الحذاء"
    ],
    "barcode": [
        "barcode", "code_barre", "code_barres", "code-barre", "ean", "ean13", "upc", "code_a_barre",
        "الباركود", "رمز الباركود", "الرمز الشريطي", "كود بار"
    ],
    "initial_stock": [
        "initial_stock", "opening_stock", "stock_initial", "stock", "quantite",
        "quantité", "qty", "qte", "qté", "stock_qty", "quantite_initiale",
        "الكمية", "المخزون الأولي", "الكمية الأولية", "رصيد أول المدة", "المخزون", "الرصيد"
    ],
    "description": [
        "description", "details", "détails", "remarque", "notes", "obs", "observations",
        "الوصف", "ملاحظات", "تفاصيل"
    ],
}

SUPPLIER_SYNONYMS: Dict[str, List[str]] = {
    "name": [
        "name", "nom", "fournisseur", "supplier", "raison_sociale", "societe", "société", "nom_fournisseur",
        "اسم المورد", "المورد", "الاسم", "اسم الشركة", "المؤسسة"
    ],
    "contact_name": [
        "contact", "contact_name", "interlocuteur", "responsable", "nom_contact", "personne_contact",
        "المسؤول", "جهة الاتصال", "اسم المسير", "الممثل"
    ],
    "phone": [
        "phone", "telephone", "téléphone", "tel", "mobile", "portable", "num_tel",
        "الهاتف", "رقم الهاتف", "الجوال", "المحمول"
    ],
    "email": [
        "email", "courriel", "mail", "e-mail",
        "البريد الإلكتروني", "الايميل", "البريد"
    ],
    "address": [
        "address", "adresse", "localisation", "siege", "siège",
        "العنوان", "المقر", "مكان التواجد"
    ],
    "origin_country": [
        "origin_country", "country", "pays", "pays_origine", "origine", "nationalite",
        "البلد", "بلد المنشأ", "الدولة"
    ],
    "payment_terms": [
        "payment_terms", "conditions_paiement", "delai_paiement", "modalites_paiement",
        "شروط الدفع", "طريقة الدفع", "آجال الدفع"
    ],
    "notes": [
        "notes", "remarque", "commentaire", "observation", "remarques",
        "ملاحظات", "تعليق"
    ],
}

CLIENT_SYNONYMS: Dict[str, List[str]] = {
    "name": [
        "name", "nom", "client", "customer", "nom_client", "raison_sociale", "nom_complet",
        "الاسم", "اسم العميل", "الزبون", "اسم الزبون", "اسم المشتري"
    ],
    "phone": [
        "phone", "telephone", "téléphone", "tel", "mobile", "portable", "num_tel",
        "الهاتف", "رقم الهاتف", "الجوال", "المحمول"
    ],
    "email": [
        "email", "courriel", "mail", "e-mail",
        "البريد الإلكتروني", "الايميل", "البريد"
    ],
    "address": [
        "address", "adresse", "domicile", "rue",
        "العنوان", "مكان الإقامة"
    ],
    "wilaya": [
        "wilaya", "province", "region", "région", "state", "departement", "code_wilaya",
        "الولاية", "المحافظة", "رقم الولاية"
    ],
    "client_type": [
        "client_type", "type_client", "type", "nature", "categorie_client",
        "نوع العميل", "فئة الزبون", "نوع الزبون", "النوع"
    ],
    "nif": [
        "nif", "identifiant_fiscal", "nif_client",
        "الرقم الجبائي", "الرقم التعريفي الجبائي", "الرمز الجبائي"
    ],
    "rc": [
        "rc", "registre_commerce", "registre_de_commerce", "num_rc",
        "السجل التجاري", "رقم السجل التجاري"
    ],
    "credit_limit": [
        "credit_limit", "plafond_credit", "limite_credit", "credit", "crédit", "solde_max",
        "حد الائتمان", "سقف الديون", "سقف الائتمان", "الحد الأقصى للدين"
    ],
    "notes": [
        "notes", "remarque", "observation", "remarques", "commentaire",
        "ملاحظات", "تعليق"
    ],
}

ENTITY_SYNONYMS = {
    "products": PRODUCT_SYNONYMS,
    "suppliers": SUPPLIER_SYNONYMS,
    "clients": CLIENT_SYNONYMS,
}


# ─────────────────────────────────────────────────────────────────────────────
# Normalization & Fuzzy Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _clean_header(s: str) -> str:
    """Normalize a header string: lowercased, stripped, accents removed, extra punctuation collapsed."""
    if not s:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[أإآ]", "ا", s)
    s = re.sub(r"ة\b", "ه", s)
    s = re.sub(r"ى\b", "ي", s)
    s_clean = re.sub(r"[\s_\-\.:;,/\\'\"]+", " ", s).strip()
    return s_clean


def match_field_to_column(header: str, entity_type: str) -> Optional[str]:
    """
    Find best matching canonical field for a raw uploaded column header.
    Returns canonical field key or None.
    """
    cleaned = _clean_header(header)
    if not cleaned:
        return None

    synonyms_map = ENTITY_SYNONYMS.get(entity_type, {})

    # Exact or alias matching
    for field, aliases in synonyms_map.items():
        if cleaned == field:
            return field
        for alias in aliases:
            cleaned_alias = _clean_header(alias)
            if cleaned == cleaned_alias:
                return field

    # Substring / partial matching
    for field, aliases in synonyms_map.items():
        for alias in aliases:
            cleaned_alias = _clean_header(alias)
            if len(cleaned_alias) >= 3 and (cleaned_alias in cleaned or cleaned in cleaned_alias):
                return field

    return None


def auto_map_columns(headers: List[str], entity_type: str) -> Dict[str, Optional[str]]:
    """
    Return dictionary mapping original raw headers to canonical system field keys.
    """
    mapping: Dict[str, Optional[str]] = {}
    assigned_fields: Set[str] = set()

    for h in headers:
        field = match_field_to_column(h, entity_type)
        if field and field not in assigned_fields:
            mapping[h] = field
            assigned_fields.add(field)
        else:
            mapping[h] = None

    return mapping


# ─────────────────────────────────────────────────────────────────────────────
# Data Parsers (Excel / CSV)
# ─────────────────────────────────────────────────────────────────────────────

def parse_file_data(raw_bytes: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    """
    Extracts headers and list of raw row dictionaries from raw CSV or XLSX bytes.
    Returns (headers, rows, error_message).
    """
    fname = filename.lower()
    if fname.endswith((".xlsx", ".xlsm", ".xltx")):
        return _parse_xlsx_data(raw_bytes)
    return _parse_csv_data(raw_bytes)


def _parse_csv_data(raw_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    text = ""
    for enc in ["utf-8-sig", "utf-8", "cp1256", "latin-1"]:
        try:
            text = raw_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue

    if not text:
        return [], [], _("Impossible de décoder le fichier CSV (encodage non supporté).")

    sample = text[:4096]
    delimiter = ","
    for sep in [";", "\t", ",", "|"]:
        if sample.count(sep) > sample.count(delimiter):
            delimiter = sep

    try:
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        rows_list = list(reader)
    except Exception as exc:
        return [], [], f"{_('Erreur de lecture CSV')}: {exc}"

    if not rows_list:
        return [], [], _("Le fichier CSV est vide.")

    raw_headers = rows_list[0]
    headers = [str(h).strip() for h in raw_headers if str(h).strip()]

    if not headers:
        return [], [], _("Aucun en-tête valide trouvé dans la première ligne du fichier.")

    rows = []
    for row in rows_list[1:]:
        if not any(str(c).strip() for c in row):
            continue
        row_dict = {}
        for idx, h in enumerate(headers):
            val = row[idx] if idx < len(row) else ""
            row_dict[h] = str(val).strip() if val is not None else ""
        rows.append(row_dict)

    return headers, rows, None


def _parse_xlsx_data(raw_bytes: bytes) -> Tuple[List[str], List[Dict[str, Any]], Optional[str]]:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True, read_only=True)
        ws = wb.active
        if not ws:
            return [], [], _("Le classeur Excel ne contient aucune feuille active.")

        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None)
        if not raw_headers:
            return [], [], _("Le fichier Excel est vide.")

        headers = [str(h).strip() for h in raw_headers if h is not None and str(h).strip()]
        if not headers:
            return [], [], _("Aucun en-tête valide trouvé dans la première ligne du fichier Excel.")

        rows = []
        for row in rows_iter:
            if not row or not any(v is not None and str(v).strip() for v in row):
                continue
            row_dict = {}
            for idx, h in enumerate(headers):
                val = row[idx] if idx < len(row) else ""
                row_dict[h] = str(val).strip() if val is not None else ""
            rows.append(row_dict)

        return headers, rows, None
    except Exception as exc:
        return [], [], f"{_('Erreur de lecture Excel')}: {exc}"


# ─────────────────────────────────────────────────────────────────────────────
# Value Parsers and Normalizers
# ─────────────────────────────────────────────────────────────────────────────

def parse_decimal_safe(val: Any, default: Decimal = Decimal("0.00")) -> Tuple[Decimal, Optional[str]]:
    if val is None or val == "":
        return default, None
    s = str(val).strip()
    s = re.sub(r"[\s\xa0DAdaدجDZDdzd€$]+", "", s)
    s = s.replace(",", ".")
    try:
        d = Decimal(s)
        if d < Decimal("0.00"):
            return default, _("Le montant ne peut pas être négatif.")
        return d, None
    except (InvalidOperation, ValueError):
        return default, _("Nombre décimal invalide: ") + str(val)


def parse_integer_safe(val: Any, default: int = 0, min_val: Optional[int] = None, max_val: Optional[int] = None) -> Tuple[int, Optional[str]]:
    if val is None or val == "":
        return default, None
    s = str(val).strip()
    try:
        num = int(round(float(s)))
        if min_val is not None and num < min_val:
            return default, f"{_('Valeur trop petite (minimum')} {min_val})"
        if max_val is not None and num > max_val:
            return default, f"{_('Valeur trop grande (maximum')} {max_val})"
        return num, None
    except (ValueError, TypeError):
        return default, _("Nombre entier invalide: ") + str(val)


def normalize_wilaya(val: Any) -> Tuple[str, Optional[str]]:
    """Maps wilaya input (code 01-58, French name, Arabic name) to standard 2-digit code."""
    if not val:
        return "", None
    s = str(val).strip()
    clean_num = re.sub(r"\D", "", s)
    if clean_num:
        try:
            num = int(clean_num)
            if 1 <= num <= 58:
                return f"{num:02d}", None
        except ValueError:
            pass

    clean_s = _clean_header(s)
    for code, name in WILAYA_CHOICES:
        if _clean_header(name) == clean_s or clean_s in _clean_header(name):
            return code, None

    AR_WILAYAS = {
        "01": "ادرار", "02": "الشلف", "03": "الاغواط", "04": "ام البواقي",
        "05": "باتنة", "06": "بجاية", "07": "بسكرة", "08": "بشار",
        "09": "البليدة", "10": "البويرة", "11": "تمنراست", "12": "تبسة",
        "13": "تلمسان", "14": "تيارت", "15": "تيزي وزو", "16": "الجزائر",
        "17": "الجلفة", "18": "جيجل", "19": "سطيف", "20": "سعيدة",
        "21": "سكيكدة", "22": "سيدي بلعباس", "23": "عنابة", "24": "قالمة",
        "25": "قسنطينة", "26": "المدية", "27": "مستغانم", "28": "المسيلة",
        "29": "معسكر", "30": "ورقلة", "31": "وهران", "32": "البيض",
        "33": "اليزي", "34": "برج بوعريريج", "35": "بومرداس", "36": "الطارف",
        "37": "تندوف", "38": "تيسمسيلت", "39": "الوادي", "40": "خنشلة",
        "41": "سوق اهراس", "42": "تيبازة", "43": "ميلة", "44": "عين الدفلى",
        "45": "النعامة", "46": "عين تموشنت", "47": "غرداية", "48": "غليزان",
        "49": "تيميمون", "50": "برج باجي مختار", "51": "اولاد جلال", "52": "بني عباس",
        "53": "عين صالح", "54": "عين قزام", "55": "تقرت", "56": "جانت",
        "57": "المغير", "58": "المنيعة",
    }
    for code, ar_name in AR_WILAYAS.items():
        if _clean_header(ar_name) in clean_s or clean_s in _clean_header(ar_name):
            return code, None

    return "", f"{_('Wilaya non reconnue')}: '{val}'"


def normalize_category(val: Any) -> str:
    """Map localized category name to CategoryChoices key."""
    if not val:
        return CategoryChoices.OTHER
    s = _clean_header(val)
    if "basket" in s or "sneaker" in s or "رياضي" in s or "حذاء رياضي" in s:
        return CategoryChoices.SNEAKERS
    if "botte" in s or "bottine" in s or "boot" in s or "حذاء عالي" in s or "بوط" in s:
        return CategoryChoices.BOOTS
    if "sandale" in s or "sandal" in s or "صندل" in s or "صندالة" in s:
        return CategoryChoices.SANDALS
    if "classique" in s or "formal" in s or "habille" in s or "كلاسيكي" in s or "رسمي" in s:
        return CategoryChoices.FORMAL
    if "sport" in s or "running" in s or "رياضة" in s:
        return CategoryChoices.SPORT
    if "enfant" in s or "kid" in s or "اطفال" in s or "ولادي" in s:
        return CategoryChoices.KIDS_SHOES
    if "pantoufle" in s or "slipper" in s or "claquette" in s or "مشب" in s or "شلاكة" in s or "خف" in s:
        return CategoryChoices.SLIPPERS
    return CategoryChoices.OTHER


def normalize_gender(val: Any) -> str:
    """Map localized gender name to GenderChoices key."""
    if not val:
        return GenderChoices.UNISEX
    s = _clean_header(val)
    if s in ["m", "homme", "men", "man", "homme / man", "رجالي", "رجال"]:
        return GenderChoices.MEN
    if s in ["f", "femme", "women", "woman", "femme / woman", "نسائي", "نساء"]:
        return GenderChoices.WOMEN
    if s in ["k", "enfant", "kids", "child", "enfant / kid", "اطفال", "اولاد", "بنات"]:
        return GenderChoices.KIDS
    return GenderChoices.UNISEX


def normalize_client_type(val: Any) -> str:
    if not val:
        return ClientTypeChoices.RETAIL
    s = _clean_header(val)
    if "gros" in s or "wholesale" in s or "pro" in s or "b2b" in s or "جملة" in s:
        return ClientTypeChoices.WHOLESALE
    return ClientTypeChoices.RETAIL


# ─────────────────────────────────────────────────────────────────────────────
# Unified Data Import Engine Class
# ─────────────────────────────────────────────────────────────────────────────

class DataImportEngine:
    """
    Enterprise Data Import & Migration Engine.
    Handles validation previews, duplicate checks, stock ledger postings,
    and bulk creation/updates across Products, Suppliers, and Clients.
    """

    def __init__(
        self,
        tenant: Tenant,
        entity_type: str,
        user=None,
        branch: Optional[Branch] = None,
        duplicate_mode: str = "skip",  # "skip" | "update" | "error"
    ):
        self.tenant = tenant
        self.entity_type = entity_type
        self.user = user
        self.branch = branch or Branch.objects.filter(tenant=tenant).first()
        self.duplicate_mode = duplicate_mode

    # ─────────────────────────────────────────────────────────────────────────
    # Validation & Dry-Run Preview
    # ─────────────────────────────────────────────────────────────────────────

    def validate_and_preview(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
        max_preview_rows: int = 50,
    ) -> Dict[str, Any]:
        """
        Performs comprehensive dry-run validation against database duplicates and schema rules.
        Returns a rich structured preview report.
        """
        if self.entity_type == "products":
            return self._validate_products(headers, rows, mapping, max_preview_rows)
        elif self.entity_type == "suppliers":
            return self._validate_suppliers(headers, rows, mapping, max_preview_rows)
        elif self.entity_type == "clients":
            return self._validate_clients(headers, rows, mapping, max_preview_rows)
        else:
            raise ValueError(f"Unknown entity type: {self.entity_type}")

    def _validate_products(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
        max_preview_rows: int,
    ) -> Dict[str, Any]:
        existing_refs = set(
            Product.objects.filter(tenant=self.tenant)
            .exclude(reference="")
            .values_list("reference", flat=True)
        )
        existing_products_keys = {
            (p.name.strip().lower(), (p.brand or "").strip().lower()): p.id
            for p in Product.objects.filter(tenant=self.tenant)
        }
        existing_barcodes = set(
            Variant.objects.filter(product__tenant=self.tenant)
            .exclude(barcode="")
            .values_list("barcode", flat=True)
        )

        preview_rows = []
        all_errors = []
        valid_count = 0
        duplicate_count = 0
        error_count = 0
        warning_count = 0

        unique_products_seen = set()
        unique_variants_seen = set()
        total_initial_stock = 0

        for idx, row in enumerate(rows, start=2):
            row_errors: List[str] = []
            row_warnings: List[str] = []
            mapped: Dict[str, Any] = {}

            for orig_header, val in row.items():
                target_field = mapping.get(orig_header)
                if target_field:
                    mapped[target_field] = val

            name = str(mapped.get("name") or "").strip()
            if not name:
                row_errors.append(str(_("Le nom du produit est obligatoire.")))

            brand = str(mapped.get("brand") or "").strip()
            reference = str(mapped.get("reference") or "").strip()
            colour = str(mapped.get("colour") or "").strip() or "N/A"

            size_val = mapped.get("size_eu")
            size_num, size_err = parse_integer_safe(size_val, default=0, min_val=15, max_val=60)
            if size_err or size_num == 0:
                row_errors.append(
                    str(_("Pointure EU invalide (doit être comprise entre 15 et 60, reçu: ")) + str(size_val) + ")"
                )
            mapped["size_eu"] = size_num

            sale_price, sp_err = parse_decimal_safe(mapped.get("sale_price"), Decimal("0.00"))
            if sp_err or sale_price <= Decimal("0.00"):
                row_errors.append(str(_("Prix de vente invalide ou manquant.")))
            mapped["sale_price"] = str(sale_price)

            purchase_price, pp_err = parse_decimal_safe(mapped.get("purchase_price"), Decimal("0.00"))
            if pp_err:
                row_warnings.append(str(_("Prix d'achat invalide, initialisé à 0.")))
            mapped["purchase_price"] = str(purchase_price)

            wholesale_price, wp_err = parse_decimal_safe(mapped.get("wholesale_price"), Decimal("0.00"))
            mapped["wholesale_price"] = str(wholesale_price)

            stock_qty, stock_err = parse_integer_safe(mapped.get("initial_stock"), default=0, min_val=0)
            if stock_err:
                row_warnings.append(str(_("Quantité de stock initial invalide, ignorée.")))
                stock_qty = 0
            mapped["initial_stock"] = stock_qty
            if not row_errors:
                total_initial_stock += stock_qty

            barcode = str(mapped.get("barcode") or "").strip()
            if barcode:
                if barcode in existing_barcodes:
                    row_warnings.append(f"{_('Code-barres déjà existant en base')}: {barcode}")
                mapped["barcode"] = barcode

            is_duplicate = False
            prod_key = (name.lower(), brand.lower())
            var_key = (name.lower(), brand.lower(), size_num, colour.lower())

            if reference and reference in existing_refs:
                is_duplicate = True
            elif prod_key in existing_products_keys and var_key in unique_variants_seen:
                is_duplicate = True

            unique_products_seen.add(prod_key)
            unique_variants_seen.add(var_key)

            if row_errors:
                status = "error"
                error_count += 1
                for err in row_errors:
                    all_errors.append({"row": idx, "field": "general", "message": err})
            elif is_duplicate:
                if self.duplicate_mode == "error":
                    status = "error"
                    error_count += 1
                    all_errors.append({"row": idx, "field": "reference", "message": str(_("Produit/variante en double."))})
                else:
                    status = "duplicate"
                    duplicate_count += 1
            elif row_warnings:
                status = "warning"
                warning_count += 1
                valid_count += 1
            else:
                status = "valid"
                valid_count += 1

            if len(preview_rows) < max_preview_rows:
                preview_rows.append({
                    "row_index": idx,
                    "status": status,
                    "errors": row_errors,
                    "warnings": row_warnings,
                    "raw": row,
                    "mapped": mapped,
                })

        return {
            "summary": {
                "total_rows": len(rows),
                "valid_rows": valid_count,
                "warning_rows": warning_count,
                "duplicate_rows": duplicate_count,
                "error_rows": error_count,
                "products_count": len(unique_products_seen),
                "variants_count": len(unique_variants_seen),
                "total_initial_stock": total_initial_stock,
            },
            "headers": headers,
            "detected_mapping": mapping,
            "preview_rows": preview_rows,
            "all_errors": all_errors[:100],
        }

    def _validate_suppliers(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
        max_preview_rows: int,
    ) -> Dict[str, Any]:
        existing_names = {
            s.name.strip().lower(): s.id
            for s in Supplier.objects.filter(tenant=self.tenant)
        }
        existing_phones = set(
            Supplier.objects.filter(tenant=self.tenant)
            .exclude(phone="")
            .values_list("phone", flat=True)
        )

        preview_rows = []
        all_errors = []
        valid_count = 0
        duplicate_count = 0
        error_count = 0
        warning_count = 0

        seen_names = set()
        seen_phones = set()

        for idx, row in enumerate(rows, start=2):
            row_errors: List[str] = []
            row_warnings: List[str] = []
            mapped: Dict[str, Any] = {}

            for orig_header, val in row.items():
                target_field = mapping.get(orig_header)
                if target_field:
                    mapped[target_field] = val

            name = str(mapped.get("name") or "").strip()
            if not name:
                row_errors.append(str(_("Le nom du fournisseur est obligatoire.")))

            phone = str(mapped.get("phone") or "").strip()
            email = str(mapped.get("email") or "").strip()
            if email and ("@" not in email or "." not in email):
                row_warnings.append(f"{_('Format email invalide')}: {email}")

            is_duplicate = False
            clean_name = name.lower()
            if clean_name in existing_names or clean_name in seen_names:
                is_duplicate = True
            elif phone and (phone in existing_phones or phone in seen_phones):
                is_duplicate = True

            seen_names.add(clean_name)
            if phone:
                seen_phones.add(phone)

            if row_errors:
                status = "error"
                error_count += 1
                for err in row_errors:
                    all_errors.append({"row": idx, "field": "name", "message": err})
            elif is_duplicate:
                if self.duplicate_mode == "error":
                    status = "error"
                    error_count += 1
                    all_errors.append({"row": idx, "field": "name", "message": str(_("Fournisseur déjà existant."))})
                else:
                    status = "duplicate"
                    duplicate_count += 1
            elif row_warnings:
                status = "warning"
                warning_count += 1
                valid_count += 1
            else:
                status = "valid"
                valid_count += 1

            if len(preview_rows) < max_preview_rows:
                preview_rows.append({
                    "row_index": idx,
                    "status": status,
                    "errors": row_errors,
                    "warnings": row_warnings,
                    "raw": row,
                    "mapped": mapped,
                })

        return {
            "summary": {
                "total_rows": len(rows),
                "valid_rows": valid_count,
                "warning_rows": warning_count,
                "duplicate_rows": duplicate_count,
                "error_rows": error_count,
            },
            "headers": headers,
            "detected_mapping": mapping,
            "preview_rows": preview_rows,
            "all_errors": all_errors[:100],
        }

    def _validate_clients(
        self,
        headers: List[str],
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
        max_preview_rows: int,
    ) -> Dict[str, Any]:
        existing_names = {
            c.name.strip().lower(): c.id
            for c in Client.objects.filter(tenant=self.tenant)
        }
        existing_phones = set(
            Client.objects.filter(tenant=self.tenant)
            .exclude(phone="")
            .values_list("phone", flat=True)
        )

        preview_rows = []
        all_errors = []
        valid_count = 0
        duplicate_count = 0
        error_count = 0
        warning_count = 0

        seen_names = set()
        seen_phones = set()

        for idx, row in enumerate(rows, start=2):
            row_errors: List[str] = []
            row_warnings: List[str] = []
            mapped: Dict[str, Any] = {}

            for orig_header, val in row.items():
                target_field = mapping.get(orig_header)
                if target_field:
                    mapped[target_field] = val

            name = str(mapped.get("name") or "").strip()
            if not name:
                row_errors.append(str(_("Le nom du client est obligatoire.")))

            phone = str(mapped.get("phone") or "").strip()
            email = str(mapped.get("email") or "").strip()
            if email and ("@" not in email or "." not in email):
                row_warnings.append(f"{_('Format email invalide')}: {email}")

            raw_wilaya = mapped.get("wilaya")
            if raw_wilaya:
                wilaya_code, w_err = normalize_wilaya(raw_wilaya)
                if w_err:
                    row_warnings.append(w_err)
                mapped["wilaya"] = wilaya_code

            credit_limit, cl_err = parse_decimal_safe(mapped.get("credit_limit"), Decimal("0.00"))
            if cl_err:
                row_warnings.append(str(_("Limite de crédit invalide, initialisée à 0.")))
            mapped["credit_limit"] = str(credit_limit)

            mapped["client_type"] = normalize_client_type(mapped.get("client_type"))

            is_duplicate = False
            clean_name = name.lower()
            if phone and (phone in existing_phones or phone in seen_phones):
                is_duplicate = True
            elif clean_name in existing_names or clean_name in seen_names:
                is_duplicate = True

            seen_names.add(clean_name)
            if phone:
                seen_phones.add(phone)

            if row_errors:
                status = "error"
                error_count += 1
                for err in row_errors:
                    all_errors.append({"row": idx, "field": "name", "message": err})
            elif is_duplicate:
                if self.duplicate_mode == "error":
                    status = "error"
                    error_count += 1
                    all_errors.append({"row": idx, "field": "name", "message": str(_("Client déjà existant."))})
                else:
                    status = "duplicate"
                    duplicate_count += 1
            elif row_warnings:
                status = "warning"
                warning_count += 1
                valid_count += 1
            else:
                status = "valid"
                valid_count += 1

            if len(preview_rows) < max_preview_rows:
                preview_rows.append({
                    "row_index": idx,
                    "status": status,
                    "errors": row_errors,
                    "warnings": row_warnings,
                    "raw": row,
                    "mapped": mapped,
                })

        return {
            "summary": {
                "total_rows": len(rows),
                "valid_rows": valid_count,
                "warning_rows": warning_count,
                "duplicate_rows": duplicate_count,
                "error_rows": error_count,
            },
            "headers": headers,
            "detected_mapping": mapping,
            "preview_rows": preview_rows,
            "all_errors": all_errors[:100],
        }

    # ─────────────────────────────────────────────────────────────────────────
    # Batch Execution Engine
    # ─────────────────────────────────────────────────────────────────────────

    def execute_import(
        self,
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
    ) -> Dict[str, Any]:
        """
        Executes the import with atomic transactions and accurate ledger postings.
        """
        if self.entity_type == "products":
            return self._execute_products(rows, mapping)
        elif self.entity_type == "suppliers":
            return self._execute_suppliers(rows, mapping)
        elif self.entity_type == "clients":
            return self._execute_clients(rows, mapping)
        else:
            raise ValueError(f"Unknown entity type: {self.entity_type}")

    def _execute_products(
        self,
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
    ) -> Dict[str, Any]:
        created_products = 0
        updated_products = 0
        created_variants = 0
        updated_variants = 0
        created_movements = 0
        skipped_count = 0
        errors: List[Dict[str, Any]] = []

        mapped_rows = []
        for idx, row in enumerate(rows, start=2):
            m: Dict[str, Any] = {}
            for orig_h, val in row.items():
                field = mapping.get(orig_h)
                if field:
                    m[field] = val
            mapped_rows.append((idx, m))

        with transaction.atomic():
            for idx, mapped in mapped_rows:
                name = str(mapped.get("name") or "").strip()
                if not name:
                    errors.append({"row": idx, "message": str(_("Nom manquant."))})
                    continue

                brand = str(mapped.get("brand") or "").strip()
                reference = str(mapped.get("reference") or "").strip()
                colour = str(mapped.get("colour") or "").strip() or "N/A"

                size_num, size_err = parse_integer_safe(mapped.get("size_eu"), default=0, min_val=15, max_val=60)
                if size_err or size_num == 0:
                    errors.append({"row": idx, "message": f"{_('Pointure invalide')}: {mapped.get('size_eu')}"})
                    continue

                sale_price, _sp_err = parse_decimal_safe(mapped.get("sale_price"), Decimal("0.00"))
                purchase_price, _pp_err = parse_decimal_safe(mapped.get("purchase_price"), Decimal("0.00"))
                wholesale_price, _wp_err = parse_decimal_safe(mapped.get("wholesale_price"), Decimal("0.00"))
                pairs_per_carton, _pc_err = parse_integer_safe(mapped.get("pairs_per_carton"), default=10, min_val=1)
                alert_threshold, _at_err = parse_integer_safe(mapped.get("alert_threshold"), default=10, min_val=0)
                stock_qty, _sq_err = parse_integer_safe(mapped.get("initial_stock"), default=0, min_val=0)
                barcode = str(mapped.get("barcode") or "").strip()
                description = str(mapped.get("description") or "").strip()

                category = normalize_category(mapped.get("category"))
                gender = normalize_gender(mapped.get("gender"))

                product = None
                if reference:
                    product = Product.objects.filter(tenant=self.tenant, reference=reference).first()
                if not product:
                    product = Product.objects.filter(tenant=self.tenant, name__iexact=name, brand__iexact=brand).first()

                is_new_product = False
                if not product:
                    product = Product.objects.create(
                        tenant=self.tenant,
                        name=name,
                        brand=brand,
                        reference=reference,
                        category=category,
                        gender=gender,
                        purchase_price=purchase_price,
                        pamp=purchase_price,
                        sale_price=sale_price,
                        wholesale_price=wholesale_price,
                        pairs_per_carton=pairs_per_carton,
                        alert_threshold=alert_threshold,
                        description=description,
                    )
                    created_products += 1
                    is_new_product = True
                elif self.duplicate_mode == "update":
                    if sale_price > 0:
                        product.sale_price = sale_price
                    if purchase_price > 0:
                        product.purchase_price = purchase_price
                    if wholesale_price > 0:
                        product.wholesale_price = wholesale_price
                    if description:
                        product.description = description
                    product.save()
                    updated_products += 1
                elif self.duplicate_mode == "error" and not is_new_product:
                    errors.append({"row": idx, "message": str(_("Produit déjà existant."))})
                    continue

                variant = Variant.objects.filter(
                    product=product,
                    size_eu=size_num,
                    colour__iexact=colour
                ).first()

                is_new_variant = False
                if not variant:
                    variant_kwargs = {
                        "tenant": self.tenant,
                        "product": product,
                        "size_eu": size_num,
                        "colour": colour,
                        "alert_threshold": 3,
                    }
                    if barcode and not Variant.objects.filter(product__tenant=self.tenant, barcode=barcode).exists():
                        variant_kwargs["barcode"] = barcode
                    variant = Variant.objects.create(**variant_kwargs)
                    created_variants += 1
                    is_new_variant = True
                elif self.duplicate_mode == "update":
                    if barcode and not Variant.objects.filter(product__tenant=self.tenant, barcode=barcode).exclude(pk=variant.pk).exists():
                        variant.barcode = barcode
                        variant.save()
                    updated_variants += 1
                else:
                    skipped_count += 1

                if stock_qty > 0 and self.branch:
                    StockMovement.objects.create(
                        tenant=self.tenant,
                        branch=self.branch,
                        variant=variant,
                        quantity_delta=stock_qty,
                        reason=MovementReasonChoices.INITIAL,
                        user=self.user,
                        notes=str(_("Import stock initial")),
                    )
                    created_movements += 1

        return {
            "success": True,
            "created_products": created_products,
            "updated_products": updated_products,
            "created_variants": created_variants,
            "updated_variants": updated_variants,
            "created_stock_movements": created_movements,
            "skipped": skipped_count,
            "errors": errors,
        }

    def _execute_suppliers(
        self,
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
    ) -> Dict[str, Any]:
        created = 0
        updated = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []

        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                mapped: Dict[str, Any] = {}
                for orig_h, val in row.items():
                    field = mapping.get(orig_h)
                    if field:
                        mapped[field] = val

                name = str(mapped.get("name") or "").strip()
                if not name:
                    errors.append({"row": idx, "message": str(_("Nom manquant."))})
                    continue

                contact_name = str(mapped.get("contact_name") or "").strip()
                phone = str(mapped.get("phone") or "").strip()
                email = str(mapped.get("email") or "").strip()
                address = str(mapped.get("address") or "").strip()
                origin_country = str(mapped.get("origin_country") or "").strip()
                payment_terms = str(mapped.get("payment_terms") or "").strip()
                notes = str(mapped.get("notes") or "").strip()

                supplier = Supplier.objects.filter(tenant=self.tenant, name__iexact=name).first()
                if not supplier and phone:
                    supplier = Supplier.objects.filter(tenant=self.tenant, phone=phone).first()

                if not supplier:
                    Supplier.objects.create(
                        tenant=self.tenant,
                        name=name,
                        contact_name=contact_name,
                        phone=phone,
                        email=email,
                        address=address,
                        origin_country=origin_country,
                        payment_terms=payment_terms,
                        notes=notes,
                    )
                    created += 1
                elif self.duplicate_mode == "update":
                    if contact_name: supplier.contact_name = contact_name
                    if phone: supplier.phone = phone
                    if email: supplier.email = email
                    if address: supplier.address = address
                    if origin_country: supplier.origin_country = origin_country
                    if payment_terms: supplier.payment_terms = payment_terms
                    if notes: supplier.notes = notes
                    supplier.save()
                    updated += 1
                elif self.duplicate_mode == "error":
                    errors.append({"row": idx, "message": str(_("Fournisseur déjà existant."))})
                else:
                    skipped += 1

        return {
            "success": True,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    def _execute_clients(
        self,
        rows: List[Dict[str, Any]],
        mapping: Dict[str, Optional[str]],
    ) -> Dict[str, Any]:
        created = 0
        updated = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []

        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                mapped: Dict[str, Any] = {}
                for orig_h, val in row.items():
                    field = mapping.get(orig_h)
                    if field:
                        mapped[field] = val

                name = str(mapped.get("name") or "").strip()
                if not name:
                    errors.append({"row": idx, "message": str(_("Nom manquant."))})
                    continue

                phone = str(mapped.get("phone") or "").strip()
                email = str(mapped.get("email") or "").strip()
                address = str(mapped.get("address") or "").strip()
                wilaya, _w_err = normalize_wilaya(mapped.get("wilaya"))
                client_type = normalize_client_type(mapped.get("client_type"))
                nif = str(mapped.get("nif") or "").strip()
                rc = str(mapped.get("rc") or "").strip()
                credit_limit, _cl_err = parse_decimal_safe(mapped.get("credit_limit"), Decimal("0.00"))
                notes = str(mapped.get("notes") or "").strip()

                client = None
                if phone:
                    client = Client.objects.filter(tenant=self.tenant, phone=phone).first()
                if not client:
                    client = Client.objects.filter(tenant=self.tenant, name__iexact=name).first()

                if not client:
                    Client.objects.create(
                        tenant=self.tenant,
                        name=name,
                        phone=phone,
                        email=email,
                        address=address,
                        wilaya=wilaya,
                        client_type=client_type,
                        nif=nif,
                        rc=rc,
                        credit_limit=credit_limit,
                        notes=notes,
                    )
                    created += 1
                elif self.duplicate_mode == "update":
                    if phone: client.phone = phone
                    if email: client.email = email
                    if address: client.address = address
                    if wilaya: client.wilaya = wilaya
                    if client_type: client.client_type = client_type
                    if nif: client.nif = nif
                    if rc: client.rc = rc
                    if credit_limit > 0: client.credit_limit = credit_limit
                    if notes: client.notes = notes
                    client.save()
                    updated += 1
                elif self.duplicate_mode == "error":
                    errors.append({"row": idx, "message": str(_("Client déjà existant."))})
                else:
                    skipped += 1

        return {
            "success": True,
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }

    # ─────────────────────────────────────────────────────────────────────────
    # Localized Sample Template Generator (XLSX & CSV)
    # ─────────────────────────────────────────────────────────────────────────

    @classmethod
    def generate_template(cls, entity_type: str, file_format: str = "xlsx", lang: str = "fr") -> Tuple[bytes, str, str]:
        """
        Builds a pre-filled, styled template with headers and sample data in FR, EN, or AR.
        Returns (content_bytes, content_type, filename).
        """
        lang = lang.lower() if lang in ["fr", "en", "ar"] else "fr"
        file_format = file_format.lower() if file_format in ["xlsx", "csv"] else "xlsx"

        templates_config = {
            "products": {
                "fr": {
                    "headers": ["Nom *", "Marque", "Référence", "Catégorie", "Genre", "Pointure *", "Couleur", "Prix Achat (DZD)", "Prix Vente (DZD) *", "Prix Gros (DZD)", "Stock Initial", "Code-barres", "Paires par Carton", "Seuil Alerte", "Description"],
                    "samples": [
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "42", "Noir", "4500", "6500", "5500", "12", "", "10", "3", "Modèle running homme confortable"],
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "43", "Noir", "4500", "6500", "5500", "15", "", "10", "3", "Modèle running homme confortable"],
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "42", "Blanc", "4500", "6500", "5500", "8", "", "10", "3", "Modèle running homme blanc"],
                        ["Mocassin Cuir Prestige", "Clarks", "CLK-MC-02", "formal", "M", "41", "Marron", "6000", "8900", "7500", "6", "", "8", "2", "Cuir véritable fait main"],
                    ],
                    "filename": "modele_import_produits_shoedz",
                },
                "en": {
                    "headers": ["Product Name *", "Brand", "Reference / SKU", "Category", "Gender", "EU Size *", "Colour", "Purchase Price (DZD)", "Sale Price (DZD) *", "Wholesale Price (DZD)", "Initial Stock", "Barcode", "Pairs per Box", "Alert Threshold", "Description"],
                    "samples": [
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "42", "Black", "4500", "6500", "5500", "12", "", "10", "3", "Comfortable men running shoes"],
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "43", "Black", "4500", "6500", "5500", "15", "", "10", "3", "Comfortable men running shoes"],
                        ["Air Max Runner", "Nike", "NK-AM-01", "sneakers", "M", "42", "White", "4500", "6500", "5500", "8", "", "10", "3", "White edition"],
                        ["Leather Loafer", "Clarks", "CLK-MC-02", "formal", "M", "41", "Brown", "6000", "8900", "7500", "6", "", "8", "2", "Handmade genuine leather"],
                    ],
                    "filename": "products_import_template_shoedz",
                },
                "ar": {
                    "headers": ["اسم المنتج *", "الماركة", "المرجع", "التصنيف", "الفئة", "المقاس *", "اللون", "سعر الشراء (دج)", "سعر البيع (دج) *", "سعر الجملة (دج)", "المخزون الأولي", "الباركود", "أزواج الكرتون", "حد التنبيه", "الوصف"],
                    "samples": [
                        ["حذاء رياضي إير ماكس", "Nike", "NK-AM-01", "sneakers", "M", "42", "أسود", "4500", "6500", "5500", "12", "", "10", "3", "حذاء رياضي مريح للرجال"],
                        ["حذاء رياضي إير ماكس", "Nike", "NK-AM-01", "sneakers", "M", "43", "أسود", "4500", "6500", "5500", "15", "", "10", "3", "حذاء رياضي مريح للرجال"],
                        ["حذاء رياضي إير ماكس", "Nike", "NK-AM-01", "sneakers", "M", "42", "أبيض", "4500", "6500", "5500", "8", "", "10", "3", "نسخة بيضاء مميزة"],
                        ["حذاء كلاسيكي جلد أصلي", "Clarks", "CLK-MC-02", "formal", "M", "41", "بني", "6000", "8900", "7500", "6", "", "8", "2", "جلد أصلي عالي الجودة"],
                    ],
                    "filename": "نموذج_استيراد_المنتجات_شوز_ديزاد",
                },
            },
            "suppliers": {
                "fr": {
                    "headers": ["Nom du Fournisseur *", "Responsable / Contact", "Téléphone", "Email", "Adresse", "Pays d'Origine", "Conditions de Paiement", "Remarques"],
                    "samples": [
                        ["Chaussures El Bahia SARL", "Karim Mansouri", "0550123456", "contact@elbahia-shoes.dz", "Zone Industrielle Oued Smar, Alger", "Algérie", "30 jours fin de mois", "Fournisseur principal cuir"],
                        ["Import Shoes International", "Mourad Belkacem", "0661987654", "import@shoestrade.com", "Route d'El Eulma, Sétif", "Italie", "Au comptant / Chèque", "Importateur officiel"],
                    ],
                    "filename": "modele_import_fournisseurs_shoedz",
                },
                "en": {
                    "headers": ["Supplier Name *", "Contact Person", "Phone", "Email", "Address", "Country", "Payment Terms", "Notes"],
                    "samples": [
                        ["El Bahia Shoes SARL", "Karim Mansouri", "0550123456", "contact@elbahia-shoes.dz", "Oued Smar Industrial Zone, Algiers", "Algeria", "30 days net", "Primary leather footwear partner"],
                        ["Import Shoes International", "Mourad Belkacem", "0661987654", "import@shoestrade.com", "El Eulma Road, Setif", "Italy", "Cheque on delivery", "Official importer"],
                    ],
                    "filename": "suppliers_import_template_shoedz",
                },
                "ar": {
                    "headers": ["اسم المورد *", "المسؤول / جهة الاتصال", "الهاتف", "البريد الإلكتروني", "العنوان", "بلد المنشأ", "شروط الدفع", "ملاحظات"],
                    "samples": [
                        ["شركة الباهية للأحذية ش.ذ.م.م", "كريم منصوري", "0550123456", "contact@elbahia-shoes.dz", "المنطقة الصناعية واد السمار، الجزائر", "الجزائر", "30 يوما بعد الاستلام", "المورد الرئيسي للأحذية الجلدية"],
                        ["مؤسسة استيراد الأحذية الدولية", "مراد بلقاسم", "0661987654", "import@shoestrade.com", "طريق العلمة، سطيف", "إيطاليا", "دفع نقدي / شيك", "مستورد رسمي"],
                    ],
                    "filename": "نموذج_استيراد_الموردين_شوز_ديزاد",
                },
            },
            "clients": {
                "fr": {
                    "headers": ["Nom du Client *", "Téléphone", "Email", "Adresse", "Wilaya (Code ou Nom)", "Type (detail / gros)", "NIF", "RC", "Plafond Crédit (DZD)", "Notes"],
                    "samples": [
                        ["Boutique Élégance Oran", "0555112233", "elegance.oran@gmail.com", "Rue Larbi Ben M'hidi", "31", "wholesale", "001234567890123", "31/00-123456B", "500000", "Client fidèle gros"],
                        ["Ahmed Benali", "0770998877", "", "Bab El Oued", "16", "retail", "", "", "0", "Client magasin détail"],
                    ],
                    "filename": "modele_import_clients_shoedz",
                },
                "en": {
                    "headers": ["Client Name *", "Phone", "Email", "Address", "Wilaya (Code or Name)", "Type (retail / wholesale)", "NIF (Tax ID)", "Commercial Register (RC)", "Credit Limit (DZD)", "Notes"],
                    "samples": [
                        ["Elegance Oran Boutique", "0555112233", "elegance.oran@gmail.com", "Larbi Ben M'hidi Street", "31", "wholesale", "001234567890123", "31/00-123456B", "500000", "Wholesale VIP partner"],
                        ["Ahmed Benali", "0770998877", "", "Bab El Oued", "16", "retail", "", "", "0", "Retail walk-in buyer"],
                    ],
                    "filename": "clients_import_template_shoedz",
                },
                "ar": {
                    "headers": ["اسم العميل *", "الهاتف", "البريد الإلكتروني", "العنوان", "الولاية (الرقم أو الاسم)", "نوع العميل (تجزئة / جملة)", "الرقم الجبائي NIF", "السجل التجاري RC", "سقف الائتمان (دج)", "ملاحظات"],
                    "samples": [
                        ["محل أناقة وهران", "0555112233", "elegance.oran@gmail.com", "شارع العربي بن مهيدي", "31", "wholesale", "001234567890123", "31/00-123456B", "500000", "زبون جملة وفي"],
                        ["أحمد بن علي", "0770998877", "", "باب الوادي", "16", "retail", "", "", "0", "زبون تجزئة"],
                    ],
                    "filename": "نموذج_استيراد_الزبائن_شوز_ديزاد",
                },
            },
        }

        conf = templates_config.get(entity_type, templates_config["products"])[lang]
        headers = conf["headers"]
        samples = conf["samples"]
        filename_base = conf["filename"]

        if file_format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in samples:
                writer.writerow(row)
            content = "\ufeff" + output.getvalue()
            return content.encode("utf-8-sig"), "text/csv; charset=utf-8", f"{filename_base}.csv"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Import Data"

        if lang == "ar":
            ws.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_thin = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center" if lang != "ar" else "right", vertical="center", wrap_text=True)
            cell.border = border_thin

        sample_font = Font(name="Calibri", size=10)
        for row in samples:
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=len(samples) + 1):
            for cell in row:
                cell.font = sample_font
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="left" if lang != "ar" else "right", vertical="center")

        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        ws.row_dimensions[1].height = 28

        out_io = io.BytesIO()
        wb.save(out_io)
        out_bytes = out_io.getvalue()

        return out_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{filename_base}.xlsx"
